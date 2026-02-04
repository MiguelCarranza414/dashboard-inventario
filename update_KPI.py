from openpyxl import load_workbook
import json
from pathlib import Path
import subprocess

excel_path = r"C:\Users\Miguel.Carranza\OneDrive - Vertiv Co\Analisis Febrero.xlsx"

# Dónde guardar los json (carpeta del script)
out_dir = Path(__file__).resolve().parent

READS = {
    "KPI.json":  ("JSON", "A16"),
    "apu1.json": ("APU1", "B21"),
    "apu2.json": ("APU2", "B21"),
    "apu3.json": ("APU3", "B22"),
}

def maybe_parse_json(value):
    # Si la celda trae un JSON como texto, lo convierte a dict/list
    if isinstance(value, str):
        s = value.strip()
        if s.startswith("{") or s.startswith("["):
            try:
                return json.loads(s)
            except json.JSONDecodeError:
                pass
    return value

wb = load_workbook(excel_path, data_only=True)

for filename, (sheet, cell) in READS.items():
    ws = wb[sheet]
    raw_value = ws[cell].value
    value = maybe_parse_json(raw_value)

    out_path = out_dir / filename
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(value, f, ensure_ascii=False, indent=2)

    print(f"✅ {filename}: {sheet}!{cell} -> OK")
#Github
REPO = Path(r"C:\KPI")
MENSAJE = "Auto update: KPI Conteos Ciclicos"

def run(cmd: list[str]) -> None:
    r = subprocess.run(cmd, cwd=REPO, text=True, capture_output=True)
    if r.returncode != 0:
        raise RuntimeError(f"Error ejecutando: {' '.join(cmd)}\n{r.stderr}")
    if r.stdout.strip():
        print(r.stdout.strip())


# 1) (opcional) ver estado
run(["git", "status"])

# 2) add (todo) o solo un archivo
# run(["git", "add", "."])
run(["git", "add", "."])  # pon la ruta relativa dentro del repo

# 3) commit (si hay cambios)
# Nota: git commit falla si "no hay nada que commitear", lo manejamos:
r = subprocess.run(["git", "diff", "--cached", "--quiet"], cwd=REPO)
if r.returncode == 0:
    print("ℹ️ No hay cambios en stage para commitear.")
else:
    run(["git", "commit", "-m", MENSAJE])
    run(["git", "push"])
    print("✅ add/commit/push completado.")
